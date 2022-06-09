# -*- coding: utf-8 -*-

# Exemplo com duas tabelas

import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import os.path

myDict = {}

# Criação da tabela padronizada
def create_table(path):
    file_excel = Workbook()
    planilha = file_excel.active
    planilha.title = "TabCropman"

    list_attribute = create_list_default()

    #tes = list_attribute[1][0]

    for col in range(1, len(list_attribute)):
       currentCell = planilha.cell(row=1, column=col, value = list_attribute[col-1][0])
       currentCell.alignment = Alignment(horizontal='center')
       ft = Font(bold = True)
       currentCell.font = ft
       #width = len(list_attribute[col-1][0])
       planilha.column_dimensions[get_column_letter(col)].width = 20

    put_values_table(planilha)

    file_excel.save(os.path.join(path, "tab_cropman.xlsx"))

def put_values_table(plan):

    list_attribute = create_list_default()

    ids = len (myDict)
    lin = 2

    for id_ in myDict.keys():
        lin+=1
        for col in range(0, len(list_attribute)):
            #print( myDict[id_][0][col][1])
            currentCell = plan.cell(row=lin, column=col+1, value = myDict[id_][0][col][1])
            currentCell.alignment = Alignment(horizontal='center')

def find_lab(table_path):
    table_name = table_path.split("\\")[-1]
    ext = table_name.split(".")[-1]

    if ext.lower()=="xls":
        #converter pra xlsx
        book = convert_xls_for_xlsx(table_path)
        convert = True
    else:
        book = load_workbook(table_path)

    table_name_ = table_name.lower()

    #1 Tentativa, pelo nome
    if "athenas" in table_name_:            return "athenas"
    elif "ibra" in table_name_:             return "ibra"
    elif "dmlab" in table_name_:            return "dmlab"
    elif "ubersolo" in table_name_:         return "ubersolo"
    elif "solos_plantas" in table_name_:    return "solos_plantas"
    elif "agrisolum" in table_name_:        return "agrisolum"
    else:
        # 2 tentativa pela planilha
        for names in book.get_sheet_names():
            sheet = book.get_sheet_by_name(names)
            n_row = sheet.max_row
            n_col = sheet.max_column
            if n_row <= 1 or n_col <=1: # Verifica se a planilha está vazia
                continue

            value = sheet.cell(row=1, column=1).value
            if value==None or value == '': #Athena, Ubersolo, Solo_Plantas, Agrisolum
                value  = sheet.cell(row=2, column=1).value
                value1 = sheet.cell(row=3, column=1).value
                value3 = sheet.cell(row=8, column=1).value
                if value3 == "Nome:": #Athenas
                    return "athenas"

                elif not value == None and not value=='':
                    return "agrisolum" #Agrisolum

                #Ubersolo, Solo_Plantas
                elif not value1== None and not value1 == '':
                    # Ubersolo, Solo_Plantas
                    #value2 = sheet.cell(row=4, column=1).value
                    if value1 == "Data":
                        return "ubersolo"
                    elif value1 == "Cod.Lab.":
                        return "solos_plantas"
                    else:
                        return "desconhecido"
                else:
                    return "desconhecido"
            else: #Ibra and DMlAb
                atrr = []
                for col in range(1,n_col):
                    atrr.append(sheet.cell(row=1, column=col).value)
                if "Gleba" in atrr:
                    return "ibra"
                elif "Classe Textural" in atrr:
                    return "dmlab"
                else:
                    return "desconhecido"

def update_values(dirp, lin, n_col, sheet, atr, flag_special=0):
    for col in range(1,n_col):
        #tt = dirp[col].strip()
        if len(dirp) > n_col: continue
        if dirp[col] == " " or dirp[col] == "": continue
        if dirp[col].strip() == atr:
            if flag_special: #Condição especial para Ubersolo
                if "K" in atr:
                    if col==20:
                        continue
                if "Na" in atr:
                    if col == 19: continue
            value = sheet.cell(row=lin, column=col).value
            break
    else:
        value = ""

    if not type(value) is str:
        #vv = value.encode('ascii', 'ignore').decode('utf-8')
        try:
            value1 = str(value)
        except:
            value1 = value
    else: value1 = value

    if value1 == 'None':
        value1 = ""

    return value1

def soma_das_bases(dirp, lin, n_col, sheet, ls_name):
    calcio, magnesio, potassio, sodio = 0,0,0,0
    calcio1_   = update_values(dirp, lin, n_col, sheet, ls_name[0])
    if not type(calcio1_) is str:
        calcio_ = str(calcio1_)
    else: calcio_ = calcio1_
    if not calcio_ == '' and not calcio_.lower()=='ns':
        calcio_ = calcio_.replace(",",".")
        calcio = float(calcio_)

    magnesio1_ = update_values(dirp, lin, n_col, sheet, ls_name[1])
    if not type(magnesio1_) is str:
        magnesio_ = str(magnesio1_)
    else: magnesio_ = magnesio1_
    if not magnesio_ == '' and not magnesio_.lower()=='ns':
        magnesio_= magnesio_.replace(",",".")
        magnesio = float(magnesio_)

    potassio1_ = update_values(dirp, lin, n_col, sheet, ls_name[2])
    if not type(potassio1_) is str:
        potassio_ = str(potassio1_)
    else: potassio_ = potassio1_
    if not potassio_ == '' and not potassio_.lower()=='ns':
        potassio_ = potassio_.replace(",",".")
        potassio = float(potassio_)

    sodio1_ = update_values(dirp, lin, n_col, sheet, ls_name[3])
    if not type(sodio1_) is str:
        sodio_ = str(sodio1_)
    else: sodio_ = sodio1_
    if not sodio_ == '' and not sodio_.lower()=='ns' :
        sodio = float(sodio_)

    return calcio + magnesio + potassio + sodio

def CTC(dirp, lin, n_col, sheet, str_aluminio, ls_name):
    aluminio1_ = update_values(dirp, lin, n_col, sheet, str_aluminio)
    if not type(aluminio1_) is str:
        aluminio_ = str(aluminio1_)
    else: aluminio_ = aluminio1_
    if aluminio_ == '' and not aluminio_.lower()=='ns':
        aluminio = 0
    else:
        aluminio_ = aluminio_.replace(",",".")
        aluminio = float(aluminio_)
    CTC_ = soma_das_bases(dirp, lin, n_col, sheet,ls_name) + aluminio

    return CTC_

def create_list_default():
    list_attribute = []
    list_attribute.append(["ID_Unico_Amonstra",""])
    list_attribute.append(["Fazenda",""])
    list_attribute.append(["Talhao",""])
    list_attribute.append(["ID_Amostra",""])
    list_attribute.append(["Data_Importacao",""])
    list_attribute.append(["Profundidade",""])
    list_attribute.append(["PH_CaCl2", ""])
    list_attribute.append(["PH_H2O", "" ])
    list_attribute.append(["MO_GDM3",""])
    list_attribute.append(["P_Res",""])
    list_attribute.append(["P_MEH",""])
    list_attribute.append(["S",""])
    list_attribute.append(["Ca",""])
    list_attribute.append(["Mg",""])
    list_attribute.append(["Na",""])
    list_attribute.append(["K",""])
    list_attribute.append(["Al",""])
    list_attribute.append(["Hal",""])
    list_attribute.append(["Soma_Bases",""])
    list_attribute.append(["CTC",""])
    list_attribute.append(["Saturacao_Bases",""])
    list_attribute.append(["Saturacao_Calcio",""])
    list_attribute.append(["Saturacao_Aluminio",""])
    list_attribute.append(["Saturacao_Magnesio",""])
    list_attribute.append(["Saturacao_Potassio",""])
    list_attribute.append(["Areia",""])
    list_attribute.append(["Argila",""])
    list_attribute.append(["Silte",""])
    list_attribute.append(["Si",""])
    list_attribute.append(["B",""])
    list_attribute.append(["Cu",""])
    list_attribute.append(["Mn",""])
    list_attribute.append(["Fe",""])
    list_attribute.append(["Zn",""])
    list_attribute.append(["Cr",""])
    list_attribute.append(["Ni",""])
    list_attribute.append(["Cd",""])
    list_attribute.append(["Pb",""])
    list_attribute.append(["Classe_Textural",""])

    return list_attribute

def sorting_depth( list_ ):
    n = 0
    list_dep = []
    depth1 = []
    depth2 = []
    depth_end = []

    while (len (list_) > n):
        list_dep.append(map(int, list_[n].split("-"))) # Transforma lista de strings em inteiros
        n += 1

    n = 0

    while (len (list_dep) > n):
        depth1.append(list_dep[n][0])
        depth2.append(list_dep[n][1])
        n += 1

    n = 1

    depth1 = sorted(set(depth1))
    depth2 = sorted(set(depth2))

    depth_f1 = str(depth1[0]) + "-" + str(depth2[0])

    while (len (depth1) > n):
        if depth_f1 in list_:
            depth_end.append(depth_f1)
        if len (depth_end) >= 2:
            break
        depth_f1 = str(depth1[n]) + "-" + str(depth2[n])
        n+=1

    return depth_end

# Padronização das tabeLas vindas do Excel
def convert_xls_for_xlsx(path_name_xls):
    #print( u"Essa planilha está no padão xls, precisa ser convertida para o padrão xlsx, convertendo...")

    # first open using xlrd
    book = xlrd.open_workbook(path_name_xls)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows+1   #bm added +1
        ncols = sheet.ncols+1   #bm added +1
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()

    for row in range(1, nrows):
        for col in range(1, ncols):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row-1, col-1) #bm added -1's

    # Teste
    '''for row in range(1, nrows):
        for col in range(1, ncols):
            print( "Row %s, Col %s Sheet %s" % (row, col, sheet1.cell(row=row, column=col).value))'''

    #print( u"Convertido!")

    return book1

    table_name = table_path.split("\\")[-1]
    ext = table_name.split(".")[-1]

    if ext.lower()=="xls":
        #converter pra xlsx
        book = convert_xls_for_xlsx(table_path)
        convert = True
    else:
        book = load_workbook(table_path)

    #print( book.sheetnames)

    for names in book.get_sheet_names():
        sheet = book.get_sheet_by_name(names)
        n_row = sheet.max_row
        n_col = sheet.max_column + 1
        if n_row <= 1 or n_col <=1: # Verifica se a planilha está vazia
            continue

        #Validando a Tabela
        B3 = sheet['B3']
        C3 = sheet['C3']

        #print( B3.value)
        #print( C3.value)

        if not u"Cod.Lab." in B3.value or not u"Descri" in C3.value:
            print( "Tabela com problemas...")
            return False

        dictonaty_atrr= {}
        #Pega os atributos
        for col in range(1,n_col):
            dictonaty_atrr[col] = sheet.cell(row=3, column=col).value

        #M4 = sheet['M4']
        N4 = sheet['N4']

        if N4.value == u"(CaCl2)":
            dictonaty_atrr[14] = u"pH (CaCl2)"

        #print( dictonaty_atrr)

       #Profundidade
        J3 = sheet['J3']
        if "Profundidade" in J3.value:
            list_ = []
            for lin in range(5,n_row + 1):
                list_.append(sheet.cell(row=lin, column=10).value)
            list_profund = sorted(set(list_)) #Ordena e elimina valores repetidos
            if len(list_profund) > 2:
                del (list_profund[2:]) #Deixa apenas as 2 profundidades mais rasas

            for lin in range(5,n_row + 1):
                prof = sheet.cell(row=lin, column=10).value
                if prof in list_profund:
                    amostra_number = sheet.cell(row=lin, column=3).value #Pega número da amostra
                    id_ = sheet.cell(row=lin, column=2).value
                    ID_UNICO_AMOSTRA = "UB_" + amostra_number + "_" + id_

                    #print( ID_UNICO_AMOSTRA)

                    myDict[ID_UNICO_AMOSTRA] = []
                    attributes_list = create_list_default()
                    myDict[ID_UNICO_AMOSTRA].append(attributes_list)

                    ls_name = (u"Ca", u"Mg", u"K", u"Na") #Nome padrões da DMLab

                    #Populando dicionario de listas
                    for index in range(0, len(attributes_list)):

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "ID_Unico_Amonstra":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = ID_UNICO_AMOSTRA
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "ID_Amostra":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = amostra_number
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Profundidade":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = prof
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Fazenda":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, "Fazenda")
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Talhao":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, u"Talhão")
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Data_Importacao":
                            data_aux = str(datetime.today()).split(" ")[-2]
                            data_inv = data_aux.split("-")
                            data = data_inv[-1] + "/" + data_inv[-2] + "/" + data_inv[-3]
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = data
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Areia":
                            areia = update_values(dictonaty_atrr, lin, n_col, sheet, "Areia total")
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(areia)

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Classe_Textural":
                            #class_tex = update_values(dictonaty_atrr, lin, n_col, sheet, "Classe Textural")
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = ""
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "PH_CaCl2":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, "pH CaCl2")
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "PH_H2O":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, "pH")
                            #myDict[ID_UNICO_AMOSTRA][0][index][1] = ""
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "MO_GDM3":
                            mat_org = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"MO"))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(mat_org)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "P_Res":
                            fosforo_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"P resina")
                            if not fosforo_ == 'ns' and not fosforo_ == "":
                                fosforo = round(((float)(fosforo_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #fosforo = round (((update_values(dictonaty_atrr, lin, n_col, sheet, u"P resina"))),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(fosforo)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "P_MEH":
                            fosforo_med_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"P meh-¹")
                            if not fosforo_med_ == 'ns' and not fosforo_med_ == "":
                                fosforo_med = round(((float)(fosforo_med_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(fosforo_med)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "S":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, u"S")
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Ca":
                            calcio_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Ca")
                            if not calcio_ == 'ns' and not calcio_ == "":
                                calcio = round((((float)(calcio_)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #calcio = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"Ca"))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(calcio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Mg":
                            mag = update_values(dictonaty_atrr, lin, n_col, sheet, u"Mg")
                            if not mag == 'ns' and not mag == "":
                                magnesio = round((((float)(mag)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #mag = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"Mg"))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(magnesio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Na":
                            sodio_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Na",1)
                            if not sodio_ == 'ns' and not sodio_ == "":
                                sodio = round((((float)(sodio_)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(sodio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "K":
                            pot = update_values(dictonaty_atrr, lin, n_col, sheet, u"K", 1)
                            if not pot == 'ns' and not pot == "":
                                potassio = round ((((float)(pot)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #pot = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"K", 1))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(potassio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Al":
                            alu = update_values(dictonaty_atrr, lin, n_col, sheet, u"Al")
                            if not alu == 'ns' and not alu == "":
                                aluminio = round ((((float)(alu)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #alu = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"Al"))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(aluminio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Hal":
                            hidro = update_values(dictonaty_atrr, lin, n_col, sheet, u"H+Al")
                            if not hidro == 'ns' and not hidro == "":
                                hidro_aluminio = round ((((float)(hidro)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #hal = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"H+Al"))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(hidro_aluminio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Si":
                            silicio_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Si")
                            if not silicio_ == 'ns' and not silicio_ == "":
                                si = round (((float)(silicio_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(si)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Soma_Bases":
                            # soma de Ca + Mg + Na + K => Ibra não tem Sódio
                            somatoria = round ((soma_das_bases(dictonaty_atrr, lin, n_col, sheet, ls_name)),2)
                            soma_ = str (somatoria)
                            soma_ = soma_.replace(".", ",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = soma_
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "CTC":
                            # soma de Ca + Mh + Na + AHl
                            ctcc = round((CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)),2)
                            CTC_ = str(ctcc)
                            CTC_ = CTC_.replace(".", ",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = CTC_
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Bases":
                            # Soma das bases * 100 / CTC
                            CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                            if CTC_ == 0:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por Zero
                                continue
                            saturacao_bases = round (((soma_das_bases(dictonaty_atrr, lin, n_col, sheet, ls_name) * 100) / CTC_), 2)

                            #Padrão Excel
                            saturacao_bases_ex = str(saturacao_bases)
                            saturacao_bases_ex = saturacao_bases_ex.replace(".",",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_bases_ex
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Calcio":
                            # Ca / CTC * 100
                            calcio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Ca")
                            if not type(calcio1_) is str:
                                calcio_ = str(calcio1_)
                            else: calcio_ = calcio1_

                            if calcio_ == '':
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                                continue
                            else:
                                calcio_ = calcio_.replace(",",".")
                            CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                            if CTC_ == 0:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por zero
                                continue
                            saturacao_calcio = round(((float(calcio_) / CTC_) * 100),2)

                            #Padrão Excel
                            saturacao_calcio_ex = str(saturacao_calcio)
                            saturacao_calcio_ex = saturacao_calcio_ex.replace(".",",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_calcio_ex
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Aluminio":
                            # (Al/(soma_bases + Al)) * 100
                            aluminio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Al")
                            if not type(aluminio1_) is str:
                                aluminio_ = str(aluminio1_)
                            else: aluminio_ = aluminio1_

                            if aluminio_ == '':
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                                continue
                            else:
                                aluminio_ = aluminio_.replace(",",".")
                            saturacao_aluminio = round(((float(aluminio_) / (soma_das_bases(dictonaty_atrr, lin, n_col, sheet, ls_name) + float(aluminio_))) * 100),2)

                            #Padrão Excel
                            saturacao_aluminio_ex = str(saturacao_aluminio)
                            saturacao_aluminio_ex = saturacao_aluminio_ex.replace(".",",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_aluminio_ex
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Potassio":
                             # K / CTC * 100
                            potassio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"K")
                            if not type(potassio1_) is str:
                                potassio_ = str(potassio1_)
                            else: potassio_ = potassio1_

                            if potassio_ == '':
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                                continue
                            else:
                                potassio_ = potassio_.replace(",",".")
                            CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                            if CTC_ == 0:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por zero
                                continue

                            saturacao_potassio = round(((float(potassio_) / CTC_) * 100),2)

                            #Padrão Excel
                            saturacao_potassio_ex = str(saturacao_potassio)
                            saturacao_potassio_ex = saturacao_potassio_ex.replace(".",",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_potassio_ex
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Magnesio":
                             # Mg / CTC * 100
                            magnesio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Mg")
                            if not type(magnesio1_) is str:
                                magnesio_ = str(magnesio1_)
                            else: magnesio_ = magnesio1_

                            if magnesio_ == '':
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                                continue
                            else:
                                magnesio_ = magnesio_.replace(",",".")
                            CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                            if CTC_ == 0:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por zero
                                continue

                            saturacao_magnesio = round (((float(magnesio_) / CTC_) * 100),2)

                            #Padrão Excel
                            saturacao_magnesio_ex = str(saturacao_magnesio)
                            saturacao_magnesio_ex = saturacao_magnesio_ex.replace(".",",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_magnesio_ex
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Argila":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, "Argila")
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Silte":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, "Silte")
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "B":
                            boro_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"B")
                            if not boro_ == 'ns' and not boro_ == "":
                                boro = round (((float)(boro_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(boro)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Cu":
                            cobre_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Cu")
                            if not cobre_ == 'ns' and not cobre_ == "":
                                cobre = round (((float)(cobre_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(cobre)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Zn":
                            zinco_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Zn")
                            if not zinco_ == 'ns' and not zinco_ == "":
                                zinco = round (((float)(zinco_)),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(zinco)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Fe":
                            ferro_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Fe")
                            if not ferro_ == 'ns' and not ferro_ == "":
                                ferro = round (((float)(ferro_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(ferro)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Cr":
                            cromo_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Cr")
                            if not cromo_ == 'ns' and not cromo_ == "":
                                cromo = round (((float)(cromo_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(cromo)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Mn":
                            manga = update_values(dictonaty_atrr, lin, n_col, sheet, u"Mn")
                            if not manga == 'ns' and not manga == "":
                                mn = round (((float)(manga)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(mn)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Ni":
                            niquel_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Ni")
                            if not niquel_ == 'ns' and not niquel_ == "":
                                niquel = round (((float)(niquel_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(niquel)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Cd":
                            cd_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Cd")
                            if not cd_ == 'ns' and not cd_ == "":
                                cd = round (((float)(cd_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(cd)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Pb":
                            chumbo_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Pb")
                            if not chumbo_ == 'ns' and not chumbo_ == "":
                                chumbo = round (((float)(chumbo_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(chumbo)
                            continue

        else:
            print( "Tabela com problemas...")
            return False
    return True
def table_solos_plantas(table_path):

    table_name = table_path.split("\\")[-1]
    ext = table_name.split(".")[-1]

    if ext.lower()=="xls":
        #converter pra xlsx
        book = convert_xls_for_xlsx(table_path)
        convert = True
    else:
        book = load_workbook(table_path)

    #print( book.sheetnames)

    for names in book.get_sheet_names():
        sheet = book.get_sheet_by_name(names)
        n_row = sheet.max_row
        n_col = sheet.max_column + 1
        if n_row <= 1 or n_col <=1: # Verifica se a planilha está vazia
            continue

        #Validando a Tabela
        A3 = sheet['A3']
        B3 = sheet['B3']

        #print( B3.value)
        #print( C3.value)

        if not u"Cod.Lab." in A3.value or not u"Descri" in B3.value:
            print( "Tabela com problemas...")
            return False

        dictonaty_atrr= {}
        #Pega os atributos
        for col in range(1,n_col):
            dictonaty_atrr[col] = sheet.cell(row=3, column=col).value

        #M4 = sheet['M4']
        K4 = sheet['K4']

        if K4.value == u"(CaCl2)":
            dictonaty_atrr[11] = u"pH (CaCl2)"

        #print( dictonaty_atrr)

       #Profundidade
        C3 = sheet['C3']
        C4 = sheet['C4']

        if "Profundidade" in C3.value or "Profundidade" in C4.value:
            list_ = []
            for lin in range(5,n_row + 1):
                list_.append(sheet.cell(row=lin, column=3).value)
            list_profund = sorted(set(list_)) #Ordena e elimina valores repetidos
            if len(list_profund) > 2:
                del (list_profund[2:]) #Deixa apenas as 2 profundidades mais rasas

            for lin in range(5,n_row + 1):
                prof = sheet.cell(row=lin, column=3).value
                if prof in list_profund:
                    amostra_number_ = sheet.cell(row=lin, column=2).value #Pega número da amostra
                    amostra_number = amostra_number_.split("-")[-1]
                    amostra_number = amostra_number.strip()
                    id_ = sheet.cell(row=lin, column=1).value
                    ID_UNICO_AMOSTRA = "SP_" + amostra_number + "_" + id_

                    #print( ID_UNICO_AMOSTRA)

                    myDict[ID_UNICO_AMOSTRA] = []
                    attributes_list = create_list_default()
                    myDict[ID_UNICO_AMOSTRA].append(attributes_list)

                    ls_name = (u"Ca", u"Mg", u"K", u"Na") #Nome padrões da DMLab

                    #Populando dicionario de listas
                    for index in range(0, len(attributes_list)):

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "ID_Unico_Amonstra":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = ID_UNICO_AMOSTRA
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "ID_Amostra":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = amostra_number
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Profundidade":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = prof
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Fazenda":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, "Fazenda")
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Talhao":
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, u"Talhão")
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Data_Importacao":
                            data_aux = str(datetime.today()).split(" ")[-2]
                            data_inv = data_aux.split("-")
                            data = data_inv[-1] + "/" + data_inv[-2] + "/" + data_inv[-3]
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = data
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Areia":
                            areia_ = update_values(dictonaty_atrr, lin, n_col, sheet, "Areia total")
                            #myDict[ID_UNICO_AMOSTRA][0][index][1] = str(areia)

                            if not areia_ == 'ns' and not areia_ == "":
                                areia = round(((float)(areia_)),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(areia)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Classe_Textural":
                            #class_tex = update_values(dictonaty_atrr, lin, n_col, sheet, "Classe Textural")
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = ""
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "PH_CaCl2":
                            ph_cacl2_ = update_values(dictonaty_atrr, lin, n_col, sheet, "pH (CaCl2)")
                            if not ph_cacl2_ == 'ns' and not ph_cacl2_ == "":
                                ph_cacl2 = round(((float)(ph_cacl2_)),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(ph_cacl2)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "PH_H2O":
                            ph_agua = update_values(dictonaty_atrr, lin, n_col, sheet, "pH")
                            if not ph_agua == 'ns' and not ph_agua == "":
                                phagua = round(((float)(ph_agua)),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(phagua)
                            #myDict[ID_UNICO_AMOSTRA][0][index][1] = ""
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "MO_GDM3":
                            mat_org = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"MO"))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(mat_org)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "P_Res":
                            fosforo_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"P resina")
                            if not fosforo_ == 'ns' and not fosforo_ == "":
                                fosforo = round(((float)(fosforo_)),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #fosforo = round (((update_values(dictonaty_atrr, lin, n_col, sheet, u"P resina"))),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(fosforo)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "P_MEH":
                            fosforo_med_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"P meh-¹")
                            if not fosforo_med_ == 'ns' and not fosforo_med_ == "":
                                fosforo_med = round(((float)(fosforo_med_)),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(fosforo_med)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "S":
                            enxofre_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"S")
                            if not enxofre_ == 'ns' and not enxofre_ == "":
                                enxofre = round((((float)(enxofre_)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(enxofre)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Ca":
                            calcio_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Ca")
                            if not calcio_ == 'ns' and not calcio_ == "":
                                calcio = round((((float)(calcio_)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #calcio = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"Ca"))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(calcio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Mg":
                            mag = update_values(dictonaty_atrr, lin, n_col, sheet, u"Mg")
                            if not mag == 'ns' and not mag == "":
                                magnesio = round((((float)(mag)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #mag = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"Mg"))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(magnesio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Na":
                            sodio_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Na",1)
                            if not sodio_ == 'ns' and not sodio_ == "":
                                sodio = round((((float)(sodio_)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(sodio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "K":
                            pot = update_values(dictonaty_atrr, lin, n_col, sheet, u"K", 1)
                            if not pot == 'ns' and not pot == "":
                                potassio = round ((((float)(pot)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #pot = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"K", 1))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(potassio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Al":
                            alu = update_values(dictonaty_atrr, lin, n_col, sheet, u"Al")
                            if not alu == 'ns' and not alu == "":
                                aluminio = round ((((float)(alu)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #alu = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"Al"))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(aluminio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Hal":
                            hidro = update_values(dictonaty_atrr, lin, n_col, sheet, u"H+Al")
                            if not hidro == 'ns' and not hidro == "":
                                hidro_aluminio = round ((((float)(hidro)) * 10),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            #hal = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"H+Al"))) * 10),2)
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(hidro_aluminio)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Si":
                            silicio_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Si")
                            if not silicio_ == 'ns' and not silicio_ == "":
                                si = round (((float)(silicio_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(si)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Soma_Bases":
                            # soma de Ca + Mg + Na + K => Ibra não tem Sódio
                            somatoria = round ((soma_das_bases(dictonaty_atrr, lin, n_col, sheet, ls_name)),2)
                            soma_ = str (somatoria)
                            soma_ = soma_.replace(".", ",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = soma_
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "CTC":
                            # soma de Ca + Mh + Na + AHl
                            ctcc = round((CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)),2)
                            CTC_ = str(ctcc)
                            CTC_ = CTC_.replace(".", ",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = CTC_
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Bases":
                            # Soma das bases * 100 / CTC
                            CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                            if CTC_ == 0:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por Zero
                                continue
                            saturacao_bases = round (((soma_das_bases(dictonaty_atrr, lin, n_col, sheet, ls_name) * 100) / CTC_), 2)

                            #Padrão Excel
                            saturacao_bases_ex = str(saturacao_bases)
                            saturacao_bases_ex = saturacao_bases_ex.replace(".",",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_bases_ex
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Calcio":
                            # Ca / CTC * 100
                            calcio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Ca")
                            if not type(calcio1_) is str:
                                calcio_ = str(calcio1_)
                            else: calcio_ = calcio1_

                            if calcio_ == '':
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                                continue
                            else:
                                calcio_ = calcio_.replace(",",".")
                            CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                            if CTC_ == 0:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por zero
                                continue
                            saturacao_calcio = round(((float(calcio_) / CTC_) * 100),2)

                            #Padrão Excel
                            saturacao_calcio_ex = str(saturacao_calcio)
                            saturacao_calcio_ex = saturacao_calcio_ex.replace(".",",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_calcio_ex
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Aluminio":
                            # (Al/(soma_bases + Al)) * 100
                            aluminio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Al")
                            if not type(aluminio1_) is str:
                                aluminio_ = str(aluminio1_)
                            else: aluminio_ = aluminio1_

                            if aluminio_ == '':
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                                continue
                            else:
                                aluminio_ = aluminio_.replace(",",".")
                            saturacao_aluminio = round(((float(aluminio_) / (soma_das_bases(dictonaty_atrr, lin, n_col, sheet, ls_name) + float(aluminio_))) * 100),2)

                            #Padrão Excel
                            saturacao_aluminio_ex = str(saturacao_aluminio)
                            saturacao_aluminio_ex = saturacao_aluminio_ex.replace(".",",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_aluminio_ex
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Potassio":
                             # K / CTC * 100
                            potassio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"K")
                            if not type(potassio1_) is str:
                                potassio_ = str(potassio1_)
                            else: potassio_ = potassio1_

                            if potassio_ == '':
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                                continue
                            else:
                                potassio_ = potassio_.replace(",",".")
                            CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                            if CTC_ == 0:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por zero
                                continue

                            saturacao_potassio = round(((float(potassio_) / CTC_) * 100),2)

                            #Padrão Excel
                            saturacao_potassio_ex = str(saturacao_potassio)
                            saturacao_potassio_ex = saturacao_potassio_ex.replace(".",",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_potassio_ex
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Magnesio":
                             # Mg / CTC * 100
                            magnesio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Mg")
                            if not type(magnesio1_) is str:
                                magnesio_ = str(magnesio1_)
                            else: magnesio_ = magnesio1_

                            if magnesio_ == '':
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                                continue
                            else:
                                magnesio_ = magnesio_.replace(",",".")
                            CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                            if CTC_ == 0:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por zero
                                continue

                            saturacao_magnesio = round (((float(magnesio_) / CTC_) * 100),2)

                            #Padrão Excel
                            saturacao_magnesio_ex = str(saturacao_magnesio)
                            saturacao_magnesio_ex = saturacao_magnesio_ex.replace(".",",")

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_magnesio_ex
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Argila":
                            argila_ = update_values(dictonaty_atrr, lin, n_col, sheet, "Argila")

                            if not argila_ == 'ns' and not argila_ == "":
                                argila = round(((float)(argila_)),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(argila)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Silte":
                            silte_ = update_values(dictonaty_atrr, lin, n_col, sheet, "Silte")
                            if not silte_ == 'ns' and not silte_ == "":
                                silte = round(((float)(silte_)),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(silte)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "B":
                            boro_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"B")
                            if not boro_ == 'ns' and not boro_ == "":
                                boro = round (((float)(boro_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(boro)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Cu":
                            cobre_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Cu")
                            if not cobre_ == 'ns' and not cobre_ == "":
                                cobre = round (((float)(cobre_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue
                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(cobre)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Zn":
                            zinco_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Zn")
                            if not zinco_ == 'ns' and not zinco_ == "":
                                zinco = round (((float)(zinco_)),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(zinco)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Fe":
                            ferro_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Fe")
                            if not ferro_ == 'ns' and not ferro_ == "":
                                ferro = round (((float)(ferro_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(ferro)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Cr":
                            cromo_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Cr")
                            if not cromo_ == 'ns' and not cromo_ == "":
                                cromo = round (((float)(cromo_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(cromo)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Mn":
                            manga = update_values(dictonaty_atrr, lin, n_col, sheet, u"Mn")
                            if not manga == 'ns' and not manga == "":
                                mn = round (((float)(manga)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(mn)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Ni":
                            niquel_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Ni")
                            if not niquel_ == 'ns' and not niquel_ == "":
                                niquel = round (((float)(niquel_)),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(niquel)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Cd":
                            cd_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Cd")
                            if not cd_ == 'ns' and not cd_ == "":
                                cd = round (((float)(cd_)) ,2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(cd)
                            continue

                        if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Pb":
                            chumbo_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Pb")
                            if not chumbo_ == 'ns' and not chumbo_ == "":
                                chumbo = round (((float)(chumbo_)),2)
                            else:
                                myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                                continue

                            myDict[ID_UNICO_AMOSTRA][0][index][1] = str(chumbo)
                            continue

        else:
            print( "Tabela com problemas...")
            return False
    return True

def table_agrisolum(table_path):

    table_name = table_path.split("\\")[-1]
    ext = table_name.split(".")[-1]

    if ext.lower()=="xls":
        #converter pra xlsx
        book = convert_xls_for_xlsx(table_path)
        convert = True
    else:
        book = load_workbook(table_path)

    #print( book.sheetnames)

    for names in book.get_sheet_names():
        sheet = book.get_sheet_by_name(names)
        n_row = sheet.max_row
        n_col = sheet.max_column + 1
        if n_row <= 1 or n_col <=1: # Verifica se a planilha está vazia
            continue

        #Validando a Tabela
        A2 = sheet['A2']
        E3 = sheet['E2']

        if not u"nº Laborat." in A2.value or not u"Descri" in E3.value:
            print( "Tabela com problemas...")
            return False

        dictonaty_atrr= {}
        #Pega os atributos
        for col in range(1,n_col):
            dictonaty_atrr[col] = sheet.cell(row=2, column=col).value

        #print( dictonaty_atrr)

       # Sem Profundidade
        for lin in range(4,n_row + 1):
            amostra_number_ = sheet.cell(row=lin, column=5).value #Pega número da amostra
            amostra_number = amostra_number_.split(" ")[-4]
            amostra_number = amostra_number.strip()
            id_ = sheet.cell(row=lin, column=1).value
            ID_UNICO_AMOSTRA = "AG_" + amostra_number + "_" + id_

            #print( ID_UNICO_AMOSTRA)

            myDict[ID_UNICO_AMOSTRA] = []
            attributes_list = create_list_default()
            myDict[ID_UNICO_AMOSTRA].append(attributes_list)

            ls_name = (u"Ca2+", u"Mg2+", u"K2+", u"Na+") #Nome padrões da DMLab

            #Populando dicionario de listas
            for index in range(0, len(attributes_list)):

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "ID_Unico_Amonstra":
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = ID_UNICO_AMOSTRA
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "ID_Amostra":
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = amostra_number_
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Profundidade":
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = ""
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Fazenda":
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, "Propriedade")
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Talhao":
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, u"Lote / Talhão")
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Data_Importacao":
                    data_aux = str(datetime.today()).split(" ")[-2]
                    data_inv = data_aux.split("-")
                    data = data_inv[-1] + "/" + data_inv[-2] + "/" + data_inv[-3]
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = data
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Areia":
                    areia = update_values(dictonaty_atrr, lin, n_col, sheet, "Areia Total")
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(areia)

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Classe_Textural":
                    #class_tex = update_values(dictonaty_atrr, lin, n_col, sheet, "Classe Textural")
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = ""
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "PH_CaCl2":
                    ph_cacl2_ = update_values(dictonaty_atrr, lin, n_col, sheet, "pH CaCl2")
                    if not ph_cacl2_ == 'ns' and not ph_cacl2_ == "":
                        ph_cacl2 = round(((float)(ph_cacl2_)),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(ph_cacl2)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "PH_H2O":
                    ph_agua = update_values(dictonaty_atrr, lin, n_col, sheet, "pH (H2O)")
                    if not ph_agua == 'ns' and not ph_agua == "":
                        phagua = round(((float)(ph_agua)),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(phagua)
                    #myDict[ID_UNICO_AMOSTRA][0][index][1] = ""
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "MO_GDM3":
                    mat_org = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"Matéria Orgãnica (MO)"))) * 10),2)
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(mat_org)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "P_Res":
                    fosforo_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"P res")
                    if not fosforo_ == 'ns' and not fosforo_ == "":
                        fosforo = round(((float)(fosforo_)),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue
                    #fosforo = round (((update_values(dictonaty_atrr, lin, n_col, sheet, u"P resina"))),2)
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(fosforo)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "P_MEH":
                    fosforo_med_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"P meh")
                    if not fosforo_med_ == 'ns' and not fosforo_med_ == "":
                        fosforo_med = round(((float)(fosforo_med_)),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(fosforo_med)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "S":
                    enxofre_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"S")
                    if not enxofre_ == 'ns' and not enxofre_ == "":
                        enxofre = round((((float)(enxofre_)) * 10),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(enxofre)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Ca":
                    calcio_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Ca2+")
                    if not calcio_ == 'ns' and not calcio_ == "":
                        calcio = round((((float)(calcio_)) * 10),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue
                    #calcio = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"Ca"))) * 10),2)
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(calcio)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Mg":
                    mag = update_values(dictonaty_atrr, lin, n_col, sheet, u"Mg2+")
                    if not mag == 'ns' and not mag == "":
                        magnesio = round((((float)(mag)) * 10),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue
                    #mag = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"Mg"))) * 10),2)
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(magnesio)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Na":
                    sodio_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Na+")
                    if not sodio_ == 'ns' and not sodio_ == "":
                        sodio = round((((float)(sodio_)) * 10),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(sodio)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "K":
                    pot = update_values(dictonaty_atrr, lin, n_col, sheet, u"K2+")
                    if not pot == 'ns' and not pot == "":
                        potassio = round ((((float)(pot)) * 10),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue
                    #pot = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"K", 1))) * 10),2)
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(potassio)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Al":
                    alu = update_values(dictonaty_atrr, lin, n_col, sheet, u"Al3+")
                    if not alu == 'ns' and not alu == "":
                        aluminio = round ((((float)(alu)) * 10),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue
                    #alu = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"Al"))) * 10),2)
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(aluminio)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Hal":
                    hidro = update_values(dictonaty_atrr, lin, n_col, sheet, u"H+ + Al3+")
                    if not hidro == 'ns' and not hidro == "":
                        hidro_aluminio = round ((((float)(hidro)) * 10),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue
                    #hal = round ((((float)(update_values(dictonaty_atrr, lin, n_col, sheet, u"H+Al"))) * 10),2)
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(hidro_aluminio)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Si":
                    silicio_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Si")
                    if not silicio_ == 'ns' and not silicio_ == "":
                        si = round (((float)(silicio_)) ,2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(si)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Soma_Bases":
                    # soma de Ca + Mg + Na + K => Ibra não tem Sódio
                    somatoria = round ((soma_das_bases(dictonaty_atrr, lin, n_col, sheet, ls_name)),2)
                    soma_ = str (somatoria)
                    soma_ = soma_.replace(".", ",")

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = soma_
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "CTC":
                    # soma de Ca + Mh + Na + AHl
                    ctcc = round((CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)),2)
                    CTC_ = str(ctcc)
                    CTC_ = CTC_.replace(".", ",")

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = CTC_
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Bases":
                    # Soma das bases * 100 / CTC
                    CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                    if CTC_ == 0:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por Zero
                        continue
                    saturacao_bases = round (((soma_das_bases(dictonaty_atrr, lin, n_col, sheet, ls_name) * 100) / CTC_), 2)

                    #Padrão Excel
                    saturacao_bases_ex = str(saturacao_bases)
                    saturacao_bases_ex = saturacao_bases_ex.replace(".",",")

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_bases_ex
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Calcio":
                    # Ca / CTC * 100
                    calcio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Ca2+")
                    if not type(calcio1_) is str:
                        calcio_ = str(calcio1_)
                    else: calcio_ = calcio1_

                    if calcio_ == '':
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                        continue
                    else:
                        calcio_ = calcio_.replace(",",".")
                    CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                    if CTC_ == 0:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por zero
                        continue
                    saturacao_calcio = round(((float(calcio_) / CTC_) * 100),2)

                    #Padrão Excel
                    saturacao_calcio_ex = str(saturacao_calcio)
                    saturacao_calcio_ex = saturacao_calcio_ex.replace(".",",")

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_calcio_ex
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Aluminio":
                    # (Al/(soma_bases + Al)) * 100
                    aluminio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Al3+")
                    if not type(aluminio1_) is str:
                        aluminio_ = str(aluminio1_)
                    else: aluminio_ = aluminio1_

                    if aluminio_ == '':
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                        continue
                    else:
                        aluminio_ = aluminio_.replace(",",".")
                    saturacao_aluminio = round(((float(aluminio_) / (soma_das_bases(dictonaty_atrr, lin, n_col, sheet, ls_name) + float(aluminio_))) * 100),2)

                    #Padrão Excel
                    saturacao_aluminio_ex = str(saturacao_aluminio)
                    saturacao_aluminio_ex = saturacao_aluminio_ex.replace(".",",")

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_aluminio_ex
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Potassio":
                     # K / CTC * 100
                    potassio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"K2+")
                    if not type(potassio1_) is str:
                        potassio_ = str(potassio1_)
                    else: potassio_ = potassio1_

                    if potassio_ == '':
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                        continue
                    else:
                        potassio_ = potassio_.replace(",",".")
                    CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                    if CTC_ == 0:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por zero
                        continue

                    saturacao_potassio = round(((float(potassio_) / CTC_) * 100),2)

                    #Padrão Excel
                    saturacao_potassio_ex = str(saturacao_potassio)
                    saturacao_potassio_ex = saturacao_potassio_ex.replace(".",",")

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_potassio_ex
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Saturacao_Magnesio":
                     # Mg / CTC * 100
                    magnesio1_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Mg2+")
                    if not type(magnesio1_) is str:
                        magnesio_ = str(magnesio1_)
                    else: magnesio_ = magnesio1_

                    if magnesio_ == '':
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = ''
                        continue
                    else:
                        magnesio_ = magnesio_.replace(",",".")
                    CTC_ = CTC(dictonaty_atrr, lin, n_col, sheet, u"Al", ls_name)
                    if CTC_ == 0:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = '' #Divisão por zero
                        continue

                    saturacao_magnesio = round (((float(magnesio_) / CTC_) * 100),2)

                    #Padrão Excel
                    saturacao_magnesio_ex = str(saturacao_magnesio)
                    saturacao_magnesio_ex = saturacao_magnesio_ex.replace(".",",")

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = saturacao_magnesio_ex
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Argila":
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, "Argila")
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Silte":
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = update_values(dictonaty_atrr, lin, n_col, sheet, "Silte")
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "B":
                    boro_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"B")
                    if not boro_ == 'ns' and not boro_ == "":
                        boro = round (((float)(boro_)) ,2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(boro)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Cu":
                    cobre_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Cu2+")
                    if not cobre_ == 'ns' and not cobre_ == "":
                        cobre = round (((float)(cobre_)) ,2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue
                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(cobre)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Zn":
                    zinco_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Zn2+")
                    if not zinco_ == 'ns' and not zinco_ == "":
                        zinco = round (((float)(zinco_)),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(zinco)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Fe":
                    ferro_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Fe2+")
                    if not ferro_ == 'ns' and not ferro_ == "":
                        ferro = round (((float)(ferro_)) ,2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(ferro)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Cr":
                    cromo_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Cr2+")
                    if not cromo_ == 'ns' and not cromo_ == "":
                        cromo = round (((float)(cromo_)) ,2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(cromo)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Mn":
                    manga = update_values(dictonaty_atrr, lin, n_col, sheet, u"Mn2+")
                    if not manga == 'ns' and not manga == "":
                        mn = round (((float)(manga)) ,2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(mn)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Ni":
                    niquel_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Ni2+")
                    if not niquel_ == 'ns' and not niquel_ == "":
                        niquel = round (((float)(niquel_)),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(niquel)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Cd":
                    cd_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Cd2+")
                    if not cd_ == 'ns' and not cd_ == "":
                        cd = round (((float)(cd_)) ,2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(cd)
                    continue

                if myDict[ID_UNICO_AMOSTRA][0][index][0] == "Pb":
                    chumbo_ = update_values(dictonaty_atrr, lin, n_col, sheet, u"Pb")
                    if not chumbo_ == 'ns' and not chumbo_ == "":
                        chumbo = round (((float)(chumbo_)),2)
                    else:
                        myDict[ID_UNICO_AMOSTRA][0][index][1] = "ns"
                        continue

                    myDict[ID_UNICO_AMOSTRA][0][index][1] = str(chumbo)
                    continue

    return True

def read_table (path_table):

    print( u"Identificando laboratório...")
    lab = find_lab(path_table)

    if "solos_plantas" in lab:
        print( u"Laboratório Solos_Plantas, carregando dados...")
        ret = table_solos_plantas(path_table)
    elif "agrisolum" in lab:
        print( u"Laboratório AgriSolum, carregando dados...")
        ret = table_agrisolum(path_table)
    else:
        print( u"Laborátorio com tabela desconhecida.")
        ret = False

    return ret

def import_table(table_):

    if not os.path.isfile(table_):
        print( u"Arquivo não encontrado.")
        print( u"Importação falhou...")
        sys.exit()
    else:
        if read_table(table_):
            path = os.path.split(table_)[0]
            create_table(path)
        else:
            print( u"Importação falhou...")
            sys.exit()

        print( u"Importação concluida com sucesso!")

if __name__ == '__main__':

    #Nomes diferentes (Testar se acha os laboratórios)
    table  = u".\\solos_plantas.xls"         #Solos__Plantas
    table  = u".\\Agrisolum.xls"             #Agrisolum

    import_table(table)